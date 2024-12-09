'''
This Module handles the interaction with the spreadsheet.
Allows for reading, updating and checking of entries
'''
__author__ = "Lukas Beck"
__date__ = "17.10.2023"

from datetime import datetime
import openpyxl


class Spreadsheet:
    'handles interaction with spreadsheets'
    # COLUMNS = {
    #     # key:          (type, column)
    #     "number":       (int, 'A'),
    #     "name":         (str, 'B'),
    #     "since":        (datetime, 'C'),
    #     "extended":     (datetime, 'D'),
    #     "email":        (str, 'E'),
    #     "collateral":   (int, 'F'),
    #     "comment":      (str, 'G'),
    #     "keys":         (int, 'H'),
    #     "rented":       (int, 'I'),
    #     "fs":           (int, 'J'),
    #     "problem":      (int, 'K'),
    #     "revoked":      (int, 'L'),
    #     "cleared":      (int, 'M'),
    #     "damage":       (int, 'N'),
    #     "extend_code":  (int, 'O'),
    #     "extend_check": (int, 'P'),
    # }
    COLUMNS = {
        # key:          column
        "number":       'A',
        "name":         'B',
        "since":        'C',
        "extended":     'D',
        "email":        'E',
        "collateral":   'F',
        "comment":      'G',
        "keys":         'H',
        "rented":       'I',
        "fs":           'J',
        "problem":      'K',
        "revoked":      'L',
        "cleared":      'M',
        "damage":       'N',
        "extend_code":  'O',
        "extend_check": 'P',
    }

    number_of_rows = 409
    number_of_cols = 16
    def __init__(self, file: str) -> None:
        self.workbook = openpyxl.load_workbook(file)
        self.worksheet = self.workbook.active
        self.file = file

    def print(self):
        '''Print the whole excel file.'''
        for row in self.worksheet.iter_rows(max_row=self.number_of_rows, max_col=self.number_of_cols, values_only=True):
            for cell in row:
                print("%s, " %cell, end =" ")
            print()
    

    def remove_entry(self, locker_nr: int) -> bool:
        '''Removes the entry for the given locker number from the spreadsheet.
        
        Args:
            locker_nr(int): Number of the locker.
        '''
        spreadsheet_entry = self.get_entry(locker_nr)
        closed_entry = {
            "number": locker_nr,
            "name": None,
            "since": None,
            "extended": None,
            "email": None,
            "collateral": None,
            "rented": None,
            "keys": spreadsheet_entry["keys"]+1,
            "fs": None
        }
        return self.update_entry(closed_entry)

    def extend_entry(self, locker_nr: int) -> bool:
        '''Extends the given locker by updating the extended date.
        
        Args:
            locker_nr(int): Number of the locker to extend.
        '''
        entry = dict({"number": locker_nr, "extended": datetime.now()})
        return self.update_entry(entry)


    def update_entry(self, updated_entry: dict) -> bool:
        '''Updates the spreadsheet according to the given entry.
        
        Args:
            updated_entry(dict): Values to update the spreadsheet with.
        Returns:
            bool: True if entry was updated, False if not.
        Raises:
            KeyError: If key is not found in the spreadsheet.
        '''
        updated = False
        # get row number of entry (is locker number +1)
        row =  str(updated_entry["number"]+1)

        # get the old entry found in the spreadsheet
        current_entry = self.get_entry(updated_entry["number"])


        for key, value in updated_entry.items():
            # check if key is a valid key
            try:
                current_entry[key]
            except KeyError:
                raise KeyError(f"Key {key} not found in current spreadsheet entry")
            
            # update value if different
            if value != current_entry[key]:
                self.worksheet[self.COLUMNS[key]+row] = value
                print("update: %10s: %30s, replacing: %s" %(key, value, current_entry[key]))
                updated = True
        
        return updated
        extended = False

        if updated_entry["name"] == current_entry["name"]:
            if updated_entry["since"] != current_entry["since"]:
                updated_entry["extended"] = updated_entry["since"]
                updated_entry["since"] = current_entry["since"]
                print(f"Contract extended with date: {updated_entry['extended']}")
                extended = True

        

        # return here if entry was removed
        if updated_entry["name"] == None:
            return updated


        if updated and not extended:            
            # change number of keys in fs-office
            if current_entry["keys"] < 1:
                raise Exception("No keys left")
            self.worksheet[self.COLUMNS["keys"][1]+row] = current_entry["keys"]-1
            print("update: %10s: %30s, replacing: %s" %("keys", current_entry["keys"]-1, current_entry["keys"])) 

            # check if someone is from fs
            is_fs = input(updated_entry["name"] + " is from fs (y/N): ")
            if is_fs.lower() == "y":
                updated_entry["fs"] = 1
                self.worksheet[self.COLUMNS["fs"][1]+row] = int(1)
                print("update: %10s: %30s, replacing: %s" %("fs", 1, current_entry["fs"]))  
        
        return updated



    def get_entry(self, locker_nr: int) -> dict:
        '''Returns a dictionary with all values for given row/entry in the spreadsheet.
        
        Args:
            locker_nr(int): Number of the locker.
        Returns:
            dict: values in given row/entry.
        '''
        entry = {}
        column = 'A'
        for value in self.worksheet[locker_nr+1]:
            try:
                key = self.worksheet[column+"1"].value
                if self.COLUMNS[key] != column:
                    raise ValueError(f"Key value pair {key}: {column} not found in COLUMN dictionary, please update it")
            except KeyError:
                raise KeyError(f"Key {key} not found in current spreadsheet entry")
            except ValueError:
                raise
            # get key from first row and value from current row
            entry[self.worksheet[column+"1"].value] = value.value
            column  = chr(ord(column)+1) # get next letter in alphabet

        return entry
    

    def check_entry(self, entry: dict, check_order=False):
        'checks if the key and the associated type of the value are those specified in columns, can also check the order'
        if check_order:
            for key_given, key_check in zip(entry, self.COLUMNS):
                if key_given != key_check:
                    raise Exception("key must be %s but is %s" %(key_check, key_given))

        for key, value in entry.items():
            # print(type(value), columns[key][0])
            if not isinstance(value, self.COLUMNS[key][0]):
                raise Exception("type must be %s but is %s; Value is %s" %(self.COLUMNS[key][0], type(value), value))
            
    # def get_table(self):
    #     '''Returns the whole locker table'''
    #     table = []
    #     for row in self.worksheet.iter_rows(max_row=self.number_of_rows, max_col=self.number_of_cols, values_only=True):
    #         table.append(row)
    #     return table
    
    def get_table(self):
        '''Returns the whole locker table as a list of dictionaries'''
        table = []
        for row in range(self.number_of_rows):
            table.append(self.get_entry(row))
        return table