'''
This Module handles the interaction with the spreadsheet.
Allows for reading, updating and checking of entries
'''
__author__ = "Lukas Beck"
__date__ = "17.10.2023"

from datetime import datetime
import openpyxl


class Spreadsheet:
    'handles interaction with spreadsheets'
    COLUMNS = {
        # key:          (type, column)
        "number":       (int, 'A'),
        "name":         (str, 'B'),
        "since":        (datetime, 'C'),
        "extended":     (datetime, 'D'),
        "email":        (str, 'E'),
        "collateral":   (int, 'F'),
        "comment":      (str, 'G'),
        "keys":         (int, 'H'),
        "rented":       (int, 'I'),
        "fs":           (int, 'J'),
        "problem":      (int, 'K'),
        "revoked":      (int, 'L'),
        "cleared":      (int, 'M'),
    }

    number_of_rows = 409
    number_of_cols = 13
    def __init__(self, file: str) -> None:
        self.workbook = openpyxl.load_workbook(file)
        self.worksheet = self.workbook.active
        self.file = file

    def print(self):
        '''Print the whole excel file.'''
        for row in self.worksheet.iter_rows(max_row=self.number_of_rows, max_col=self.number_of_cols, values_only=True):
            for cell in row:
                print("%s, " %cell, end =" ")
            print()
    

    def remove_entry(self, locker_nr: int) -> bool:
        '''Removes the entry for the given locker number from the spreadsheet.
        
        Args:
            locker_nr(int): Number of the locker.
        '''
        spreadsheet_entry = self.get_entry(locker_nr)
        closed_entry = {
            "number": locker_nr,
            "name": None,
            "since": None,
            "extended": None,
            "email": None,
            "collateral": None,
            "rented": None,
            "keys": spreadsheet_entry["keys"]+1,
            "fs": None
        }
        return self.update_entry(closed_entry)


    def update_entry(self, new_entry: dict) -> bool:
        '''Updates the spreadsheet according to the given entry.
        
        Args:
            new_entry(dict): Values to update the spreadsheet with.
        '''
        row =  str(new_entry["number"]+1)

        spreadsheet_entry = self.get_entry(new_entry["number"])
        updated = False
        extended = False

        if new_entry["name"] == spreadsheet_entry["name"]:
            if new_entry["since"] != spreadsheet_entry["since"]:
                new_entry["extended"] = new_entry["since"]
                new_entry["since"] = spreadsheet_entry["since"]
                print(f"Contract extended with date: {new_entry['extended']}")
                extended = True

        for key, value in new_entry.items():
            if value != spreadsheet_entry[key]: # only update value if different
                self.worksheet[self.COLUMNS[key][1]+row] = value
                print("update: %10s: %30s, replacing: %s" %(key, value, spreadsheet_entry[key]))
                updated = True

        # return here if entry was removed
        if new_entry["name"] == None:
            return updated


        if updated and not extended:            
            # change number of keys in fs-office
            if spreadsheet_entry["keys"] < 1:
                raise Exception("No keys left")
            self.worksheet[self.COLUMNS["keys"][1]+row] = spreadsheet_entry["keys"]-1
            print("update: %10s: %30s, replacing: %s" %("keys", spreadsheet_entry["keys"]-1, spreadsheet_entry["keys"])) 

            # check if someone is from fs
            is_fs = input(new_entry["name"] + " is from fs (y/N): ")
            if is_fs.lower() == "y":
                new_entry["fs"] = 1
                self.worksheet[self.COLUMNS["fs"][1]+row] = int(1)
                print("update: %10s: %30s, replacing: %s" %("fs", 1, spreadsheet_entry["fs"]))  
        
        return updated


    def get_entry(self, locker_nr: int) -> dict:
        '''Returns a dictionary with all values for given row/entry in the spreadsheet.
        
        Args:
            locker_nr(int): locker_nr(int): Number of the locker.
        Returns:
            dict: values in given row/entry.
        '''
        entry = {}
        column = 'A'
        for value in self.worksheet[locker_nr+1]:
            entry[self.worksheet[column+"1"].value] = value.value
            column  = chr(ord(column)+1) # get next letter in alphabet
        return entry
    

    def check_entry(self, entry: dict, check_order=False):
        'checks if the key and the associated type of the value are those specified in columns, can also check the order'
        if check_order:
            for key_given, key_check in zip(entry, self.COLUMNS):
                if key_given != key_check:
                    raise Exception("key must be %s but is %s" %(key_check, key_given))

        for key, value in entry.items():
            # print(type(value), columns[key][0])
            if not isinstance(value, self.COLUMNS[key][0]):
                raise Exception("type must be %s but is %s; Value is %s" %(self.COLUMNS[key][0], type(value), value))
            
    def get_table(self):
        '''Returns the whole locker table'''
        table = []
        for row in self.worksheet.iter_rows(max_row=self.number_of_rows, max_col=self.number_of_cols, values_only=True):
            table.append(row)
        return table